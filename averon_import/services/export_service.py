from __future__ import annotations

from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

from averon_import.core.constants import COLUMN_BY_KEY
from averon_import.core.normalizers import as_excel_number


class ExcelExportService:
    def export(
        self,
        rows: list[dict],
        columns: list[str],
        output_path: Path,
        sheet_name: str = "Спецификация",
        include_headers: bool = True,
        only_exportable: bool = True,
    ) -> Path:
        valid_columns = [column for column in columns if column in COLUMN_BY_KEY]
        if not valid_columns:
            raise ValueError("Не выбрано ни одного столбца для экспорта")

        workbook = Workbook()
        sheet = workbook.active
        sheet.title = (sheet_name or "Спецификация")[:31]
        sheet.sheet_view.showGridLines = False

        header_fill = PatternFill("solid", fgColor="1F2937")
        header_font = Font(color="FFFFFF", bold=True, size=11)
        thin = Side(style="thin", color="D1D5DB")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)

        row_cursor = 1
        if include_headers:
            for column_index, key in enumerate(valid_columns, start=1):
                cell = sheet.cell(row_cursor, column_index, COLUMN_BY_KEY[key]["title"])
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(
                    horizontal="center", vertical="center", wrap_text=True
                )
                cell.border = border
            sheet.row_dimensions[row_cursor].height = 34
            row_cursor += 1

        exported_count = 0
        for row in rows:
            if only_exportable and row.get("row_type") in {
                "section",
                "system",
                "note",
                "skip",
            }:
                continue
            if row.get("selected") is False:
                continue
            values = []
            for key in valid_columns:
                value = row.get(key, "")
                if key in {"quantity", "mass"}:
                    value = as_excel_number(value)
                values.append(value)
            if not any(value not in (None, "") for value in values):
                continue

            for column_index, value in enumerate(values, start=1):
                cell = sheet.cell(row_cursor, column_index, value)
                cell.border = border
                cell.alignment = Alignment(vertical="top", wrap_text=True)
                if row.get("status") in {"review", "unrecognized"}:
                    cell.fill = PatternFill("solid", fgColor="FEF3C7")
            row_cursor += 1
            exported_count += 1

        for column_index, key in enumerate(valid_columns, start=1):
            letter = get_column_letter(column_index)
            configured = COLUMN_BY_KEY[key]["width"]
            sheet.column_dimensions[letter].width = min(max(configured, 10), 60)

        sheet.freeze_panes = "A2" if include_headers else None
        sheet.auto_filter.ref = sheet.dimensions
        if include_headers and exported_count > 0:
            table_ref = f"A1:{get_column_letter(len(valid_columns))}{row_cursor - 1}"
            table = Table(displayName="AveronImportTable", ref=table_ref)
            table.tableStyleInfo = TableStyleInfo(
                name="TableStyleMedium2",
                showFirstColumn=False,
                showLastColumn=False,
                showRowStripes=True,
                showColumnStripes=False,
            )
            sheet.add_table(table)

        metadata = workbook.create_sheet("Сведения")
        metadata.sheet_view.showGridLines = False
        metadata.append(["Параметр", "Значение"])
        metadata.append(["Программа", "Averon Import"])
        metadata.append(["Дата экспорта", datetime.now().strftime("%d.%m.%Y %H:%M")])
        metadata.append(["Экспортировано строк", exported_count])
        metadata.append(["Разработчик", "Андриянов Степан Владимирович - НВСС"])
        metadata.column_dimensions["A"].width = 26
        metadata.column_dimensions["B"].width = 56
        for cell in metadata[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.border = border
        for row in metadata.iter_rows(min_row=2):
            for cell in row:
                cell.border = border
                cell.alignment = Alignment(vertical="top", wrap_text=True)

        output_path.parent.mkdir(parents=True, exist_ok=True)
        workbook.save(output_path)
        return output_path
