from openpyxl import load_workbook

from averon_import.services.export_service import ExcelExportService


def test_export_selected_columns(tmp_path):
    rows = [
        {
            "name": "Воздуховод 500×500",
            "type_mark": "ГОСТ 14918-2020",
            "unit": "м",
            "quantity": "12,5",
            "row_type": "item",
            "status": "verified",
            "selected": True,
        },
        {
            "name": "Раздел",
            "row_type": "section",
            "status": "recognized",
            "selected": True,
        },
    ]
    target = tmp_path / "result.xlsx"
    ExcelExportService().export(rows, ["name", "unit", "quantity"], target)
    workbook = load_workbook(target)
    sheet = workbook["Спецификация"]
    assert sheet.max_row == 2
    assert sheet["A2"].value == "Воздуховод 500×500"
    assert sheet["C2"].value == 12.5
    assert workbook["Сведения"]["B5"].value == "Андриянов Степан Владимирович - НВСС"
